from __future__ import annotations

import csv
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook


BACKEND_ROOT = Path(__file__).resolve().parents[1]
HTML_PATH = BACKEND_ROOT / 'data' / 'example2.html'
XLSX_PATH = BACKEND_ROOT / 'data' / '2025_season.xlsx'
CSV_PATH = BACKEND_ROOT / 'data' / '2025_season.csv'

GROUP_COLUMN_MAPPING = [
    ('第一组', '第一阶段 轮次 1'),
    ('第二组', '第一阶段 轮次 2'),
    ('第三组', '第一阶段 轮次 3'),
    ('第四组', '第一阶段 轮次 4'),
    ('第五组', '第一阶段 轮次 5'),
]
GROUP_TEXT_PATTERN = re.compile(r'^//\s*第([一二三四五六])组\s*$')
GROUP_ORDER = {
    '一': 0,
    '二': 1,
    '三': 2,
    '四': 3,
    '五': 4,
    '六': 5,
}


@dataclass(frozen=True)
class ContestantEntry:
    name: str
    series: str


@dataclass(frozen=True)
class ContestantVoteEntry:
    name: str
    series: str
    votes: int


class ArenaVotesParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.current_group_index: int | None = None
        self.group_depth = 0
        self.groups: list[list[ContestantVoteEntry]] = [[] for _ in GROUP_COLUMN_MAPPING]
        self.in_name = False
        self.in_series = False
        self.in_votes = False
        self.current_name_parts: list[str] = []
        self.current_series_parts: list[str] = []
        self.current_vote_parts: list[str] = []
        self.current_name = ''
        self.current_series = ''
        self.pending_group_text_parts: list[str] = []

    def flush_pending_group_text(self) -> None:
        if not self.pending_group_text_parts:
            return

        pending_text = normalize_text(''.join(self.pending_group_text_parts))
        self.pending_group_text_parts = []
        match = GROUP_TEXT_PATTERN.match(pending_text)
        if not match:
            return

        self.current_group_index = GROUP_ORDER[match.group(1)]
        self.group_depth = 0
        self.current_name = ''
        self.current_series = ''
        self.current_name_parts = []
        self.current_series_parts = []
        self.current_vote_parts = []
        self.in_name = False
        self.in_series = False
        self.in_votes = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.flush_pending_group_text()
        if self.current_group_index is None:
            return

        attrs_dict = dict(attrs)
        class_name = attrs_dict.get('class', '') or ''

        if tag == 'div' and 'resultsArenaGroup' in class_name:
            if self.group_depth == 0:
                self.group_depth = 1
                return
            self.group_depth += 1
            return

        if self.group_depth > 0 and tag == 'div':
            self.group_depth += 1

        if self.group_depth == 0:
            return

        if tag == 'p' and 'contestantName' in class_name:
            self.in_name = True
            self.current_name_parts = []
        elif tag == 'p' and 'contestantSeries' in class_name:
            self.in_series = True
            self.current_series_parts = []
        elif tag == 'h3' and 'contestantVotes' in class_name:
            self.in_votes = True
            self.current_vote_parts = []

    def handle_data(self, data: str) -> None:
        if self.group_depth == 0 and not self.in_name and not self.in_series and not self.in_votes:
            if data.strip():
                self.pending_group_text_parts.append(data)
            return

        if self.group_depth == 0:
            return

        if self.in_name:
            self.current_name_parts.append(data)
        if self.in_series:
            self.current_series_parts.append(data)
        if self.in_votes:
            self.current_vote_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if self.group_depth == 0:
            return

        if tag == 'p' and self.in_name:
            self.in_name = False
            self.current_name = normalize_text(''.join(self.current_name_parts))
            self.current_name_parts = []
            return

        if tag == 'p' and self.in_series:
            self.in_series = False
            self.current_series = normalize_text(''.join(self.current_series_parts))
            self.current_series_parts = []
            return

        if tag == 'h3' and self.in_votes:
            self.in_votes = False
            vote_text = normalize_text(''.join(self.current_vote_parts)).replace(',', '')
            self.current_vote_parts = []
            if self.current_group_index is not None and self.current_name and self.current_series and vote_text:
                self.groups[self.current_group_index].append(
                    ContestantVoteEntry(
                        name=self.current_name,
                        series=self.current_series,
                        votes=int(vote_text),
                    )
                )
            self.current_name = ''
            self.current_series = ''
            return

        if tag == 'div':
            self.group_depth -= 1


def normalize_text(value: str) -> str:
    return ' '.join(str(value).replace('\xa0', ' ').split())


def parse_group_votes(path: Path) -> list[list[ContestantVoteEntry]]:
    parser = ArenaVotesParser()
    parser.feed(path.read_text(encoding='utf-8'))
    return parser.groups


def backup_file(path: Path) -> Path:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = path.with_name(f'{path.stem}.backup_{timestamp}{path.suffix}')
    shutil.copy2(path, backup_path)
    return backup_path


def build_header_index(worksheet) -> dict[str, int]:
    header_index: dict[str, int] = {}
    for column_idx, cell in enumerate(worksheet[1], start=1):
        header_name = normalize_text(cell.value)
        if header_name:
            header_index[header_name] = column_idx
    return header_index


def build_row_index(worksheet) -> tuple[dict[ContestantEntry, int], list[ContestantEntry]]:
    row_index: dict[ContestantEntry, int] = {}
    duplicate_entries: list[ContestantEntry] = []

    for row_idx in range(2, worksheet.max_row + 1):
        name = normalize_text(worksheet.cell(row=row_idx, column=2).value)
        series = normalize_text(worksheet.cell(row=row_idx, column=3).value)
        if not name or not series:
            continue

        entry = ContestantEntry(name=name, series=series)
        if entry in row_index:
            duplicate_entries.append(entry)
            continue

        row_index[entry] = row_idx

    return row_index, duplicate_entries


def backfill_votes(worksheet, groups: list[list[ContestantVoteEntry]]) -> tuple[int, dict[str, list[ContestantEntry]]]:
    header_index = build_header_index(worksheet)
    row_index, duplicate_entries = build_row_index(worksheet)
    if duplicate_entries:
        duplicate_text = ', '.join(f'{entry.name} @ {entry.series}' for entry in duplicate_entries)
        raise ValueError(f'2025_season.xlsx 存在重复角色行: {duplicate_text}')

    missing_columns = [column_name for _, column_name in GROUP_COLUMN_MAPPING if column_name not in header_index]
    if missing_columns:
        raise ValueError(f'2025_season.xlsx 缺少列: {", ".join(missing_columns)}')

    if len(groups) != len(GROUP_COLUMN_MAPPING):
        raise ValueError(f'{HTML_PATH.name} 分组数量异常: 期望 {len(GROUP_COLUMN_MAPPING)} 组，实际 {len(groups)} 组')

    updated_cells = 0
    unmatched_by_column: dict[str, list[ContestantEntry]] = {}

    for group_entries, (_, column_name) in zip(groups, GROUP_COLUMN_MAPPING):
        column_idx = header_index[column_name]
        unmatched_entries: list[ContestantEntry] = []

        for group_entry in group_entries:
            entry = ContestantEntry(name=group_entry.name, series=group_entry.series)
            row_idx = row_index.get(entry)
            if row_idx is None:
                unmatched_entries.append(entry)
                continue

            worksheet.cell(row=row_idx, column=column_idx).value = group_entry.votes
            updated_cells += 1

        if unmatched_entries:
            unmatched_by_column[column_name] = unmatched_entries

    return updated_cells, unmatched_by_column


def export_worksheet_to_csv(worksheet, csv_path: Path) -> None:
    with csv_path.open('w', encoding='utf-8-sig', newline='') as file_obj:
        writer = csv.writer(file_obj)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(list(row))


def main() -> int:
    groups = parse_group_votes(HTML_PATH)
    backup_path = backup_file(XLSX_PATH)

    workbook = load_workbook(XLSX_PATH)
    worksheet = workbook.active

    updated_cells, unmatched_by_column = backfill_votes(worksheet, groups)

    workbook.save(XLSX_PATH)
    export_worksheet_to_csv(worksheet, CSV_PATH)

    print(f'已备份原文件: {backup_path}')
    print(f'已回填分组数: {len(groups)}')
    print(f'已更新票数字段数: {updated_cells}')
    print(f'已写回 Excel: {XLSX_PATH}')
    print(f'已同步 CSV: {CSV_PATH}')

    if unmatched_by_column:
        print('\n以下页面角色未在 2025_season.xlsx 中找到对应行:')
        for column_name, entries in unmatched_by_column.items():
            print(f'[{column_name}]')
            for entry in entries:
                print(f'{entry.name} @ {entry.series}')
    else:
        print('所有页面角色都已按 2025_season.xlsx 原序号成功回填')

    return 0


if __name__ == '__main__':
    raise SystemExit(main())


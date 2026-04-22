from __future__ import annotations

import csv
import shutil
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook

from fill_single_vote_cell_config import INPUT_BLOCKS, TARGET_COLUMN, TARGET_ENTRIES


BACKEND_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = BACKEND_ROOT / 'data' / '2025_season.xlsx'
CSV_PATH = BACKEND_ROOT / 'data' / '2025_season.csv'

INPUT_MODE_DIRECT = 'direct'
INPUT_MODE_HTML_DUEL = 'html_duel'
INPUT_MODE_HTML_RANKED = 'html_ranked'
PICK_RULE_LAST_NON_EMPTY = 'last_non_empty'
PICK_RULE_MAX = 'max'


@dataclass(frozen=True)
class VoteEntry:
    name: str
    series: str
    value: int | float


@dataclass(frozen=True)
class ResolvedVoteEntry:
    name: str
    series: str
    value: int | float | str


class DuelResultsHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_name = False
        self.in_series = False
        self.in_votes = False
        self.current_name_parts: list[str] = []
        self.current_series_parts: list[str] = []
        self.current_vote_parts: list[str] = []
        self.current_name = ''
        self.current_series = ''
        self.entries: list[VoteEntry] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = dict(attrs)
        class_name = attrs_dict.get('class', '') or ''

        if tag == 'p' and 'contestantName' in class_name:
            self.in_name = True
            self.current_name_parts = []
        elif tag == 'p' and 'contestantSeries' in class_name:
            self.in_series = True
            self.current_series_parts = []
        elif tag == 'h3' and 'contestantVotes' in class_name:
            self.in_votes = True
            self.current_vote_parts = []

    def handle_data(self, data: str) -> None:
        if self.in_name:
            self.current_name_parts.append(data)
        if self.in_series:
            self.current_series_parts.append(data)
        if self.in_votes:
            self.current_vote_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == 'p' and self.in_name:
            self.in_name = False
            self.current_name = normalize_text(''.join(self.current_name_parts))
            self.current_name_parts = []
            return

        if tag == 'p' and self.in_series:
            self.in_series = False
            self.current_series = normalize_text(''.join(self.current_series_parts))
            self.current_series_parts = []
            return

        if tag == 'h3' and self.in_votes:
            self.in_votes = False
            vote_text = normalize_text(''.join(self.current_vote_parts)).replace(',', '')
            self.current_vote_parts = []
            if self.current_name and self.current_series and vote_text:
                self.entries.append(
                    VoteEntry(
                        name=self.current_name,
                        series=self.current_series,
                        value=parse_vote_value(vote_text),
                    )
                )
                self.current_name = ''
                self.current_series = ''


class RankedChoiceHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_header_name = False
        self.in_header_series = False
        self.in_vote = False
        self.current_name_parts: list[str] = []
        self.current_series_parts: list[str] = []
        self.current_vote_parts: list[str] = []
        self.headers: list[tuple[str, str]] = []
        self.rows: list[list[int | float | None]] = []
        self.current_row: list[int | float | None] | None = None
        self.current_cell_has_vote = False
        self.in_thead_row = False
        self.in_body = False
        self.row_depth = 0
        self.cell_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = dict(attrs)
        class_name = attrs_dict.get('class', '') or ''

        if tag == 'tbody':
            self.in_body = True
            return

        if tag == 'tr' and 'resultsTableRow' in class_name:
            self.in_thead_row = True
            return

        if tag == 'tr' and self.in_body and not self.in_thead_row:
            self.current_row = []
            self.row_depth = 1
            return

        if self.current_row is not None and tag == 'td':
            self.current_row.append(None)
            self.cell_depth = 1
            self.current_cell_has_vote = False
            return

        if self.current_row is not None and self.cell_depth > 0 and tag == 'td':
            self.cell_depth += 1
            return

        if tag == 'p' and 'contestantName' in class_name:
            self.in_header_name = True
            self.current_name_parts = []
        elif tag == 'p' and 'contestantSeries' in class_name:
            self.in_header_series = True
            self.current_series_parts = []
        elif tag == 'h3' and 'contestantVotes' in class_name:
            self.in_vote = True
            self.current_vote_parts = []

    def handle_data(self, data: str) -> None:
        if self.in_header_name:
            self.current_name_parts.append(data)
        if self.in_header_series:
            self.current_series_parts.append(data)
        if self.in_vote:
            self.current_vote_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == 'p' and self.in_header_name:
            self.in_header_name = False
            self.current_name_parts = [normalize_text(''.join(self.current_name_parts))]
            return

        if tag == 'p' and self.in_header_series:
            self.in_header_series = False
            name = self.current_name_parts[0] if self.current_name_parts else ''
            series = normalize_text(''.join(self.current_series_parts))
            if name and series:
                self.headers.append((name, series))
            self.current_series_parts = []
            self.current_name_parts = []
            return

        if tag == 'h3' and self.in_vote:
            self.in_vote = False
            vote_text = normalize_text(''.join(self.current_vote_parts)).replace(',', '')
            self.current_vote_parts = []
            if self.current_row is not None and self.current_row and vote_text:
                self.current_row[-1] = parse_vote_value(vote_text)
                self.current_cell_has_vote = True
            return

        if tag == 'td' and self.current_row is not None and self.cell_depth > 0:
            self.cell_depth -= 1
            if self.cell_depth == 0:
                self.current_cell_has_vote = False
            return

        if tag == 'tr' and self.current_row is not None:
            self.rows.append(self.current_row)
            self.current_row = None
            self.row_depth = 0
            return

        if tag == 'tr' and self.in_thead_row:
            self.in_thead_row = False
            return

        if tag == 'tbody':
            self.in_body = False


def normalize_text(value: object) -> str:
    if value is None:
        return ''
    return ' '.join(str(value).replace('\xa0', ' ').split())


def format_vote_value(value: int | float) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def parse_vote_value(value_text: str) -> int | float:
    normalized = value_text.strip()
    if '.' in normalized:
        return float(normalized)
    return int(normalized)


def merge_duplicate_entries(entries: list[VoteEntry]) -> list[ResolvedVoteEntry]:
    grouped_values: dict[tuple[str, str], list[int | float]] = {}
    order: list[tuple[str, str]] = []

    for entry in entries:
        key = (entry.name, entry.series)
        if key not in grouped_values:
            grouped_values[key] = []
            order.append(key)
        grouped_values[key].append(entry.value)

    merged_entries: list[ResolvedVoteEntry] = []
    for name, series in order:
        values = grouped_values[(name, series)]
        if len(values) == 1:
            merged_value: int | float | str = values[0]
        else:
            merged_value = '/'.join(format_vote_value(value) for value in values)
        merged_entries.append(ResolvedVoteEntry(name=name, series=series, value=merged_value))

    return merged_entries


def resolve_block_entries(block: dict[str, object]) -> list[VoteEntry]:
    mode = str(block['mode'])

    if mode == INPUT_MODE_DIRECT:
        entries_data = block['entries']
        if not isinstance(entries_data, list):
            raise ValueError('direct block 的 entries 必须是 list')
        return [VoteEntry(name=name, series=series, value=value) for name, series, value in entries_data]

    if mode == INPUT_MODE_HTML_DUEL:
        html_text = str(block['html'])
        return parse_duel_html(html_text)

    if mode == INPUT_MODE_HTML_RANKED:
        html_text = str(block['html'])
        pick_rule = str(block.get('pick_rule', PICK_RULE_LAST_NON_EMPTY))
        return parse_ranked_html(html_text, pick_rule)

    raise ValueError(f'不支持的 block mode: {mode}')


def resolve_target_entries() -> list[tuple[str, str, int | float | str]]:
    if not INPUT_BLOCKS:
        raise ValueError('INPUT_BLOCKS 不能为空')

    all_entries: list[VoteEntry] = []
    for block in INPUT_BLOCKS:
        all_entries.extend(resolve_block_entries(block))
    return [(entry.name, entry.series, entry.value) for entry in merge_duplicate_entries(all_entries)]


def parse_duel_html(html_text: str) -> list[VoteEntry]:
    parser = DuelResultsHtmlParser()
    parser.feed(html_text)
    if not parser.entries:
        raise ValueError('未从 HTML 中解析到任何对战票数')
    return parser.entries


def parse_ranked_html(html_text: str, pick_rule: str) -> list[VoteEntry]:
    parser = RankedChoiceHtmlParser()
    parser.feed(html_text)

    if not parser.headers:
        raise ValueError('未从 HTML 中解析到排序复选制表头')
    if not parser.rows:
        raise ValueError('未从 HTML 中解析到排序复选制票数行')
    if len(parser.headers) != len(parser.rows[0]):
        raise ValueError(f'排序复选制表头与首行列数不一致: headers={len(parser.headers)}, row={len(parser.rows[0])}')

    entries: list[VoteEntry] = []
    for column_index, (name, series) in enumerate(parser.headers):
        column_values = [row[column_index] for row in parser.rows if column_index < len(row) and row[column_index] is not None]
        if not column_values:
            raise ValueError(f'排序复选制角色没有任何票数: {name} @ {series}')

        if pick_rule == PICK_RULE_LAST_NON_EMPTY:
            picked_value = column_values[-1]
        elif pick_rule == PICK_RULE_MAX:
            picked_value = max(column_values)
        else:
            raise ValueError(f'不支持的 PICK_RULE: {pick_rule}')

        entries.append(VoteEntry(name=name, series=series, value=picked_value))

    return entries


def backup_file(path: Path) -> Path:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = path.with_name(f'{path.stem}.backup_{timestamp}{path.suffix}')
    shutil.copy2(path, backup_path)
    return backup_path


def build_header_index(worksheet) -> dict[str, int]:
    header_index: dict[str, int] = {}
    for column_idx, cell in enumerate(worksheet[1], start=1):
        header_name = normalize_text(cell.value)
        if header_name:
            header_index[header_name] = column_idx
    return header_index


def find_target_row(worksheet, target_name: str, target_series: str) -> int:
    matched_rows: list[int] = []

    for row_idx in range(2, worksheet.max_row + 1):
        name = normalize_text(worksheet.cell(row=row_idx, column=2).value)
        series = normalize_text(worksheet.cell(row=row_idx, column=3).value)
        if name == target_name and series == target_series:
            matched_rows.append(row_idx)

    if not matched_rows:
        raise ValueError(f'未找到角色: {target_name} @ {target_series}')

    if len(matched_rows) > 1:
        raise ValueError(f'找到重复角色行: {target_name} @ {target_series} -> {matched_rows}')

    return matched_rows[0]


def clear_non_target_values(worksheet, column_idx: int, target_entries: list[tuple[str, str, int | float | str]]) -> list[tuple[str, str, object]]:
    target_keys = {(target_name, target_series) for target_name, target_series, _ in target_entries}
    cleared_entries: list[tuple[str, str, object]] = []

    for row_idx in range(2, worksheet.max_row + 1):
        name = normalize_text(worksheet.cell(row=row_idx, column=2).value)
        series = normalize_text(worksheet.cell(row=row_idx, column=3).value)
        if not name or not series:
            continue
        if (name, series) in target_keys:
            continue

        cell = worksheet.cell(row=row_idx, column=column_idx)
        if cell.value in (None, ''):
            continue

        cleared_entries.append((name, series, cell.value))
        cell.value = None

    return cleared_entries


def backfill_target_entries(worksheet, target_column: str, target_entries: list[tuple[str, str, int | float | str]]) -> tuple[list[tuple[str, str, object]], list[tuple[str, str, object, int | float | str]]]:
    header_index = build_header_index(worksheet)
    column_idx = header_index.get(target_column)
    if column_idx is None:
        raise ValueError(f'未找到轮次列: {target_column}')

    cleared_entries = clear_non_target_values(worksheet, column_idx, target_entries)

    updated_entries: list[tuple[str, str, object, int | float | str]] = []
    for target_name, target_series, target_value in target_entries:
        row_idx = find_target_row(worksheet, target_name, target_series)
        old_value = worksheet.cell(row=row_idx, column=column_idx).value
        worksheet.cell(row=row_idx, column=column_idx).value = target_value
        updated_entries.append((target_name, target_series, old_value, target_value))

    return cleared_entries, updated_entries


def export_worksheet_to_csv(worksheet, csv_path: Path) -> None:
    with csv_path.open('w', encoding='utf-8-sig', newline='') as file_obj:
        writer = csv.writer(file_obj)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(list(row))


def main() -> int:
    target_entries = resolve_target_entries()
    backup_path = backup_file(XLSX_PATH)

    workbook = load_workbook(XLSX_PATH)
    worksheet = workbook.active

    cleared_entries, updated_entries = backfill_target_entries(worksheet, TARGET_COLUMN, target_entries)

    workbook.save(XLSX_PATH)
    export_worksheet_to_csv(worksheet, CSV_PATH)

    print(f'已备份原文件: {backup_path}')
    print(f'输入块数: {len(INPUT_BLOCKS)}')
    print(f'已写入轮次: {TARGET_COLUMN}')
    print(f'已清空其他角色数: {len(cleared_entries)}')
    for target_name, target_series, old_value in cleared_entries:
        print(f'清空 {target_name} @ {target_series}: {old_value} -> None')
    print(f'已更新角色数: {len(updated_entries)}')
    for target_name, target_series, old_value, target_value in updated_entries:
        print(f'{target_name} @ {target_series}: {old_value} -> {target_value}')
    print(f'已写回 Excel: {XLSX_PATH}')
    print(f'已同步 CSV: {CSV_PATH}')

    return 0


if __name__ == '__main__':
    raise SystemExit(main())

import os
import sys

import pandas as pd
from openpyxl import load_workbook


def _normalize_cell_value(value):
    if pd.isna(value):
        return None
    return value


def csv_to_excel(csv_path, excel_path=None):
    """将 CSV 数据写回现有 Excel，同时保留原有样式。"""
    try:
        if not os.path.exists(csv_path):
            print(f"错误：文件不存在 - {csv_path}")
            return None

        if excel_path is None:
            excel_path = csv_path.rsplit('.', 1)[0] + '.xlsx'

        if not os.path.exists(excel_path):
            print(f"错误：Excel 模板不存在 - {excel_path}")
            return None

        print(f"正在读取 CSV 文件: {csv_path}")
        df = pd.read_csv(csv_path)

        print(f"正在加载 Excel 模板: {excel_path}")
        workbook = load_workbook(excel_path)
        worksheet = workbook.active

        csv_headers = list(df.columns)
        csv_rows = df.values.tolist()
        target_row_count = len(csv_rows) + 1
        target_col_count = len(csv_headers)
        max_row = max(worksheet.max_row, target_row_count)
        max_col = max(worksheet.max_column, target_col_count)

        for row_index in range(1, max_row + 1):
            for col_index in range(1, max_col + 1):
                if row_index == 1 and col_index <= target_col_count:
                    worksheet.cell(row=row_index, column=col_index, value=csv_headers[col_index - 1])
                    continue

                data_row_index = row_index - 2
                data_col_index = col_index - 1
                if 0 <= data_row_index < len(csv_rows) and data_col_index < target_col_count:
                    value = _normalize_cell_value(csv_rows[data_row_index][data_col_index])
                    worksheet.cell(row=row_index, column=col_index, value=value)
                else:
                    worksheet.cell(row=row_index, column=col_index, value=None)

        workbook.save(excel_path)

        print(f"\n✓ 成功将 {csv_path} 写回到 {excel_path}")
        print("\n已保留原 Excel 的样式、行高、列宽等格式信息")
        print(f"\n数据预览 (前5行):\n{df.head()}")
        print(f"\n总共 {len(df)} 行, {len(df.columns)} 列")

        return df

    except Exception as error:
        print(f"转换出错: {error}")
        return None


if __name__ == "__main__":
    print("=" * 60)
    print("CSV 写回 Excel 工具（保留格式）")
    print("=" * 60)

    if len(sys.argv) < 2:
        print("用法: python csv_to_excel.py <csv_path> [excel_path]")
        sys.exit(1)

    csv_path = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else None
    csv_to_excel(csv_path, excel_path)
